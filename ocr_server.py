import os
import re
import tempfile
import shutil
import json
import requests
from flask import Flask, request, jsonify
from flask_cors import CORS
from dotenv import load_dotenv

# --- FIX FOR PADDLEPADDLE C++ BUG (ConvertPirAttribute2RuntimeAttribute) ---
os.environ["FLAGS_enable_pir_api"] = "0"
os.environ["FLAGS_enable_pir_in_executor"] = "0"
os.environ["FLAGS_use_mkldnn"] = "0"
os.environ["paddle_disable_pir"] = "1"

load_dotenv()  # Load .env file from python-backend/

app = Flask(__name__)
CORS(app)

# Groq API — fast cloud LLM (LPU hardware, ~1s response)
GROQ_API_KEY = os.environ.get('GROQ_API_KEY', '')
GROQ_MODEL = os.environ.get('GROQ_MODEL', 'llama-3.3-70b-versatile')
GROQ_URL = 'https://api.groq.com/openai/v1/chat/completions'

# ─── Initialize PaddleOCR (English) ───────────────────────────────────────────
try:
    print("Initializing PaddleOCR (English)...")
    from paddleocr import PaddleOCR
    ocr_en = PaddleOCR(use_angle_cls=True, lang='en', enable_mkldnn=False)
    print("PaddleOCR (English) initialized successfully.")
except Exception as e:
    print(f"Warning: PaddleOCR (English) failed: {e}")
    ocr_en = None

# ─── Initialize PaddleOCR (Hindi) ─────────────────────────────────────────────
try:
    print("Initializing PaddleOCR (Hindi)...")
    from paddleocr import PaddleOCR as PaddleOCR_Hindi
    ocr_hi = PaddleOCR_Hindi(use_angle_cls=True, lang='hi', enable_mkldnn=False)
    print("PaddleOCR (Hindi) initialized successfully.")
except Exception as e:
    print(f"Warning: PaddleOCR (Hindi) failed: {e}")
    ocr_hi = None

# ─── Initialize ChromaDB and RAG ──────────────────────────────────────────────
import time
try:
    print("Initializing RAG (ChromaDB & SentenceTransformers)...")
    import chromadb
    from sentence_transformers import SentenceTransformer
    chroma_client = chromadb.Client()
    embedder = SentenceTransformer('all-MiniLM-L6-v2')
    RAG_ENABLED = True
    print("RAG initialized successfully.")
except Exception as e:
    print(f"Warning: RAG initialization failed: {e}")
    RAG_ENABLED = False


# ─── Helper: Enhance image for better OCR ─────────────────────────────────────
def enhance_image(image_path, output_path):
    """Enhance image contrast and denoise for better OCR accuracy"""
    try:
        from PIL import Image, ImageFilter, ImageEnhance
        img = Image.open(image_path).convert('RGB')
        # Increase sharpness
        img = ImageEnhance.Sharpness(img).enhance(2.0)
        # Increase contrast
        img = ImageEnhance.Contrast(img).enhance(1.5)
        # Apply slight sharpening filter
        img = img.filter(ImageFilter.SHARPEN)
        img.save(output_path)
        return output_path
    except Exception as e:
        print(f"Image enhancement failed (using original): {e}")
        return image_path


def run_ocr(image_path, lang='en'):
    """Run PaddleOCR on image, auto-detect Hindi or English"""
    ocr = ocr_hi if lang == 'hi' and ocr_hi else ocr_en
    if not ocr:
        return ""
    result = ocr.ocr(image_path)
    if not result or not result[0]:
        return ""

    boxes = []
    for line in result[0]:
        try:
            box = line[0]
            # Handle different PaddleOCR return structures safely
            if isinstance(line[1], (tuple, list)):
                text_val = line[1][0] if len(line[1]) > 0 else ""
            else:
                text_val = str(line[1]) if line[1] else ""
                
            ys = [p[1] for p in box]
            xs = [p[0] for p in box]
            boxes.append({
                'top': min(ys),
                'bottom': max(ys),
                'x0': min(xs),
                'x1': max(xs),
                'text': text_val
            })
        except Exception as e:
            print(f"Skipping malformed OCR line: {e}")
            continue

    # Group OCR boxes by approximate vertical line using bounding box overlap
    boxes.sort(key=lambda w: w['top'])
    lines = []
    current_line = []
    current_top = None
    current_bottom = None
    
    for w in boxes:
        if current_top is None:
            current_line.append(w)
            current_top = w['top']
            current_bottom = w['bottom']
        else:
            overlap = max(0, min(current_bottom, w['bottom']) - max(current_top, w['top']))
            word_height = w['bottom'] - w['top']
            # If overlap is > 40% of the word's height, consider it the same line
            if word_height > 0 and (overlap / word_height) > 0.4:
                current_line.append(w)
                current_top = min(current_top, w['top'])
                current_bottom = max(current_bottom, w['bottom'])
            else:
                lines.append(current_line)
                current_line = [w]
                current_top = w['top']
                current_bottom = w['bottom']
    if current_line:
        lines.append(current_line)

    text_result = ""
    for line_words in lines:
        line_words = sorted(line_words, key=lambda w: w['x0'])
        line_text = " ".join([w['text'] for w in line_words])
        text_result += line_text + "\n"

    return text_result.strip()


# ─── Helper: Convert pdfplumber table to Markdown ─────────────────────────────
def table_to_markdown(table):
    """Convert a pdfplumber table (list of lists) to Markdown table format"""
    if not table or not table[0]:
        return ""
    rows = []
    for i, row in enumerate(table):
        cells = [str(cell or '').strip() for cell in row]
        rows.append("| " + " | ".join(cells) + " |")
        if i == 0:
            rows.append("| " + " | ".join(["---"] * len(row)) + " |")
    return "\n".join(rows)


# ─── Helper: Reconstruct PDF text with formatting ─────────────────────────────
def reconstruct_page_text(words, tables_boxes):
    """Group words into lines, maintain horizontal layout spacing, and detect headings/bold"""
    if not words:
        return ""

    # Filter out words that are inside tables to prevent duplicates
    filtered_words = []
    for w in words:
        in_table = False
        for box in tables_boxes:
            # box is (x0, top, x1, bottom)
            if box[0] <= w['x0'] <= box[2] and box[1] <= w['top'] <= box[3]:
                in_table = True
                break
        if not in_table:
            filtered_words.append(w)
            
    if not filtered_words:
        return ""

    # Gather all font sizes to compute median (normal body text size)
    all_sizes = [w.get('size', 10) for w in filtered_words if w.get('size')]
    if all_sizes:
        all_sizes.sort()
        median_size = all_sizes[len(all_sizes) // 2]
    else:
        median_size = 10

    # Group words by approximate vertical line using bounding box overlap
    filtered_words.sort(key=lambda w: w['top'])
    lines = []
    current_line = []
    current_top = None
    current_bottom = None
    
    for w in filtered_words:
        if current_top is None:
            current_line.append(w)
            current_top = w['top']
            current_bottom = w['bottom']
        else:
            overlap = max(0, min(current_bottom, w['bottom']) - max(current_top, w['top']))
            word_height = w['bottom'] - w['top']
            # If overlap is > 40% of the word's height, consider it the same line
            if word_height > 0 and (overlap / word_height) > 0.4:
                current_line.append(w)
                current_top = min(current_top, w['top'])
                current_bottom = max(current_bottom, w['bottom'])
            else:
                lines.append(current_line)
                current_line = [w]
                current_top = w['top']
                current_bottom = w['bottom']
    if current_line:
        lines.append(current_line)

    result_lines = []
    for line_words in lines:
        line_words = sorted(line_words, key=lambda w: w['x0'])
        
        # Reconstruct line maintaining horizontal spacing
        line_text = ""
        last_x1 = line_words[0]['x0'] if line_words else 0
        
        for w in line_words:
            gap = w['x0'] - last_x1
            # Estimate average space width based on font size (usually ~0.25 to 0.3 of font size)
            space_width = max(w.get('size', median_size) * 0.3, 2.0)
            
            if gap > space_width * 1.5:
                # Insert multiple spaces to preserve visual layout (columns, tabs, indents)
                num_spaces = int(gap / space_width)
                line_text += " " * num_spaces
            elif line_text and gap > 0:
                line_text += " "
                
            line_text += w['text']
            last_x1 = w['x1']
            
        # Remove (cid:xxx) artifacts
        line_text = re.sub(r'\(cid:\d+\)', '', line_text).rstrip()
        if not line_text.strip():
            continue

        avg_size = sum(w.get('size', median_size) for w in line_words) / len(line_words)
        is_bold = any('Bold' in w.get('fontname', '') or 'bold' in w.get('fontname', '') for w in line_words)

        if avg_size >= median_size * 1.6:          # Big heading (H2)
            result_lines.append(f"\n## {line_text.strip()}\n")
        elif avg_size >= median_size * 1.3:        # Sub-heading (H3)
            result_lines.append(f"\n### {line_text.strip()}\n")
        elif is_bold and avg_size >= median_size * 1.1:
            result_lines.append(f"**{line_text.strip()}**")
        else:
            result_lines.append(line_text)

    return '\n'.join(result_lines)


# ─── /api/extract-image ───────────────────────────────────────────────────────
@app.route('/api/extract-image', methods=['POST'])
def extract_image():
    if not ocr_en:
        return jsonify({"error": "PaddleOCR is not initialized on the server"}), 500

    if 'file' not in request.files:
        return jsonify({"error": "No file part in the request"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    lang = request.form.get('lang', 'en')  # Accept lang param: 'en' or 'hi'

    temp_dir = tempfile.mkdtemp()
    try:
        raw_path = os.path.join(temp_dir, "raw_image.png")
        enhanced_path = os.path.join(temp_dir, "enhanced_image.png")
        file.save(raw_path)

        # Enhance image before OCR
        final_path = enhance_image(raw_path, enhanced_path)

        print(f"Running PaddleOCR ({lang}) on image: {file.filename}")
        extracted_text = run_ocr(final_path, lang=lang)

        return jsonify({
            "success": True,
            "text": extracted_text,
            "confidence": 95
        })
    except Exception as e:
        print(f"Error extracting image: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


# ─── /api/extract-pdf ─────────────────────────────────────────────────────────
@app.route('/api/extract-pdf', methods=['POST'])
def extract_pdf():
    if 'file' not in request.files:
        return jsonify({"error": "No file part in the request"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    lang = request.form.get('lang', 'en')  # 'en' or 'hi'

    temp_dir = tempfile.mkdtemp()
    temp_path = os.path.join(temp_dir, "input.pdf")
    file.save(temp_path)

    try:
        print(f"Extracting from PDF: {file.filename} (lang={lang})")
        import pdfplumber

        all_sections = []
        is_scanned = False

        with pdfplumber.open(temp_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                page_content = []

                # ── 1. Extract Tables ──────────────────────────────────────
                tables = page.extract_tables()
                table_bboxes = []
                if tables:
                    raw_tables = page.find_tables()
                    for tbl_obj, tbl_data in zip(raw_tables, tables):
                        md_table = table_to_markdown(tbl_data)
                        if md_table:
                            page_content.append(f"\n{md_table}\n")
                            table_bboxes.append(tbl_obj.bbox)

                # ── 2. Extract Words (non-table area) ─────────────────────
                words = page.extract_words(
                    x_tolerance=2,
                    y_tolerance=3,
                    keep_blank_chars=False,
                    use_text_flow=True,
                    extra_attrs=['fontname', 'size']
                )

                if words:
                    page_text = reconstruct_page_text(words, table_bboxes)
                    if page_text.strip():
                        page_content.insert(0, page_text)  # Text before tables
                    elif not tables:
                        is_scanned = True
                else:
                    is_scanned = True

                # ── 3. Fallback: Scanned page → OCR ───────────────────────
                if is_scanned and not tables:
                    print(f"  Page {page_num + 1}: scanned, running OCR...")
                    if ocr_en:
                        img = page.to_image(resolution=200).original
                        img_path = os.path.join(temp_dir, f"page_{page_num}.png")
                        enh_path = os.path.join(temp_dir, f"page_{page_num}_enh.png")
                        img.save(img_path)
                        final_img = enhance_image(img_path, enh_path)
                        ocr_text = run_ocr(final_img, lang=lang)
                        if ocr_text:
                            page_content = [ocr_text]

                if page_content:
                    combined = '\n'.join(page_content).strip()
                    if combined:
                        all_sections.append(f"## Page {page_num + 1}\n\n{combined}")
                is_scanned = False  # Reset for next page

        full_text = "\n\n---\n\n".join(all_sections)

        if not full_text.strip():
            raise Exception("No text could be extracted from the PDF.")

        print(f"Successfully extracted {len(full_text)} characters from PDF.")
        return jsonify({
            "success": True,
            "text": full_text,
            "confidence": 95
        })

    except Exception as e:
        print(f"Error extracting PDF: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


# ─── /api/ai/summarize ────────────────────────────────────────────────────────
@app.route('/api/ai/summarize', methods=['POST'])
def ai_summarize():
    data = request.get_json()
    text = data.get('text', '')
    style = data.get('style', 'bullet')  # bullet | paragraph | key-points
    length = data.get('length', 'medium')  # short | medium | long

    if not text.strip():
        return jsonify({"error": "No text provided"}), 400

    length_map = {
        'short': '2-3 sentences',
        'medium': '1 concise paragraph (4-6 sentences)',
        'long': '2-3 detailed paragraphs'
    }
    style_map = {
        'bullet': 'as clear bullet points',
        'paragraph': 'as a paragraph',
        'key-points': 'as numbered key points'
    }

    prompt = f"""Summarize this document {length_map.get(length, '1 paragraph')} {style_map.get(style, 'as bullet points')}.
Focus on main ideas. If text is in Hindi, summarize in Hindi.

Document:
{text[:3000]}

Summary:"""

    result = call_ollama(prompt, max_tokens=200)
    return jsonify({"success": True, "summary": result})


# ─── /api/ai/qa ───────────────────────────────────────────────────────────────
@app.route('/api/ai/qa', methods=['POST'])
def ai_qa():
    data = request.get_json()
    text = data.get('text', '')
    question = data.get('question', '')

    if not text.strip() or not question.strip():
        return jsonify({"error": "Text and question required"}), 400

    prompt = f"""Answer this question based ONLY on the document. If not found, say so.
If document is in Hindi, answer in Hindi.

Document: {text[:3000]}

Question: {question}
Answer:"""

    result = call_ollama(prompt, max_tokens=150)
    return jsonify({"success": True, "answer": result})


# ─── /api/ai/entities ─────────────────────────────────────────────────────────
@app.route('/api/ai/entities', methods=['POST'])
def ai_entities():
    data = request.get_json()
    text = data.get('text', '')

    if not text.strip():
        return jsonify({"error": "No text provided"}), 400

    prompt = f"""Extract key entities from this text. List: person names, organizations, locations, dates, emails, phones, important keywords.
Format: Category: value (one per line)
If Hindi text, respond in Hindi.

Text: {text[:2000]}"""

    result = call_ollama(prompt, max_tokens=150)
    return jsonify({"success": True, "entities": result})


# ─── /api/ai/improve ──────────────────────────────────────────────────────────
@app.route('/api/ai/improve', methods=['POST'])
def ai_improve():
    data = request.get_json()
    text = data.get('text', '')
    style = data.get('style', 'professional')

    if not text.strip():
        return jsonify({"error": "No text provided"}), 400

    style_map = {
        'professional': 'in a professional, formal style',
        'casual': 'in a casual, friendly style',
        'academic': 'in an academic, scholarly style',
        'simple': 'in simple, easy-to-understand language'
    }

    prompt = f"""Rewrite the following text {style_map.get(style, 'professionally')}. 
Improve grammar, clarity, and flow while preserving the original meaning.
If the text is in Hindi, rewrite in Hindi.

Text:
{text[:4000]}

Rewritten:"""

    result = call_ollama(prompt)
    return jsonify({"success": True, "improved": result})


# ─── /api/ai/translate ────────────────────────────────────────────────────────
@app.route('/api/ai/translate', methods=['POST'])
def ai_translate():
    data = request.get_json()
    text = data.get('text', '')
    target_lang = data.get('target_lang', 'Hindi')

    if not text.strip():
        return jsonify({"error": "No text provided"}), 400

    prompt = f"""Translate to {target_lang}. Reply with translation ONLY, no explanation.

Text: {text[:2000]}

Translation:"""

    result = call_ollama(prompt, max_tokens=200)
    return jsonify({"success": True, "translated": result})


# ─── Groq AI Helper ───────────────────────────────────────────────────────
def call_ollama(prompt, max_tokens=128):
    """Call Groq API (fast cloud LPU) — replaces slow Ollama VPS"""
    if not GROQ_API_KEY:
        return "[Error] GROQ_API_KEY not set in .env"
    try:
        response = requests.post(
            GROQ_URL,
            headers={
                'Authorization': f'Bearer {GROQ_API_KEY}',
                'Content-Type': 'application/json'
            },
            json={
                'model': GROQ_MODEL,
                'messages': [{'role': 'user', 'content': prompt}],
                'max_tokens': max_tokens,
                'temperature': 0.2
            },
            timeout=30   # Groq is fast, 30s is more than enough
        )
        response.raise_for_status()
        return response.json()['choices'][0]['message']['content'].strip()
    except Exception as e:
        print(f"Groq API error: {e}")
        return f"[Error] {str(e)}"


# ─── /api/ai/detect-type ─────────────────────────────────────────────────────
@app.route('/api/ai/detect-type', methods=['POST'])
def ai_detect_type():
    data = request.get_json()
    text = data.get('text', '')[:1500]
    if not text.strip():
        return jsonify({"error": "No text provided"}), 400

    prompt = f"""Identify the document type from this text. Reply with ONLY one of these words:
resume, invoice, id_card, legal, academic, medical, news_article, letter, question_paper, other

Text:
{text}

Document type:"""

    doc_type = call_ollama(prompt, max_tokens=10).lower().strip()
    # Clean up response to just get the type
    valid_types = ['resume', 'invoice', 'id_card', 'legal', 'academic', 'medical', 'news_article', 'letter', 'question_paper', 'other']
    detected = next((t for t in valid_types if t in doc_type), 'other')
    return jsonify({"success": True, "type": detected})


# ─── /api/ai/extract-structured ──────────────────────────────────────────────
@app.route('/api/ai/extract-structured', methods=['POST'])
def ai_extract_structured():
    data = request.get_json()
    text = data.get('text', '')[:1500]
    doc_type = data.get('doc_type', 'resume')
    if not text.strip():
        return jsonify({"error": "No text provided"}), 400

    prompts = {
        'resume': f"Extract from this resume: name, email, phone, skills (list), experience (list), education. Return as JSON.\n\nResume:\n{text}\n\nJSON:",
        'invoice': f"Extract from this invoice: invoice_no, date, vendor, total_amount, gst_no, items (list). Return as JSON.\n\nInvoice:\n{text}\n\nJSON:",
        'id_card': f"Extract from this ID card: name, dob, id_number, address, issuing_authority. Return as JSON.\n\nID Card:\n{text}\n\nJSON:",
        'medical': f"Extract from this medical doc: patient_name, date, diagnosis, medicines, doctor_name. Return as JSON.\n\nDoc:\n{text}\n\nJSON:",
        'other': f"Extract all key information as JSON key-value pairs.\n\nDocument:\n{text}\n\nJSON:"
    }
    prompt = prompts.get(doc_type, prompts['other'])
    result = call_ollama(prompt, max_tokens=400)

    # Try to parse JSON from result
    try:
        json_match = re.search(r'\{.*\}', result, re.DOTALL)
        if json_match:
            parsed = json.loads(json_match.group())
            return jsonify({"success": True, "data": parsed, "raw": result})
    except:
        pass
    return jsonify({"success": True, "data": None, "raw": result})


# ─── /api/ai/redact ───────────────────────────────────────────────────────────
@app.route('/api/ai/redact', methods=['POST'])
def ai_redact():
    data = request.get_json()
    text = data.get('text', '')
    if not text.strip():
        return jsonify({"error": "No text provided"}), 400

    # Regex-based PII redaction (fast, no AI needed)
    redacted = text
    # Phone numbers (Indian + international)
    redacted = re.sub(r'(\+91[-\s]?)?[6-9]\d{9}', '[PHONE]', redacted)
    redacted = re.sub(r'\+?\d[\d\s\-()]{8,}\d', '[PHONE]', redacted)
    # Email addresses
    redacted = re.sub(r'[\w.+-]+@[\w-]+\.[\w.]+', '[EMAIL]', redacted)
    # Aadhar number (12 digits)
    redacted = re.sub(r'\b\d{4}\s?\d{4}\s?\d{4}\b', '[AADHAR]', redacted)
    # PAN card
    redacted = re.sub(r'\b[A-Z]{5}[0-9]{4}[A-Z]\b', '[PAN]', redacted)
    # Credit/Debit card numbers
    redacted = re.sub(r'\b\d{4}[\s-]?\d{4}[\s-]?\d{4}[\s-]?\d{4}\b', '[CARD]', redacted)
    # GST number
    redacted = re.sub(r'\b\d{2}[A-Z]{5}\d{4}[A-Z]{1}[A-Z\d]{1}[Z]{1}[A-Z\d]{1}\b', '[GST]', redacted)

    changes = text.count('[') - redacted.count('[')  # rough count
    pii_count = sum(1 for m in re.finditer(r'\[(PHONE|EMAIL|AADHAR|PAN|CARD|GST)\]', redacted))

    return jsonify({"success": True, "redacted_text": redacted, "pii_found": pii_count})


# ─── /api/ai/compare ──────────────────────────────────────────────────────────
@app.route('/api/ai/compare', methods=['POST'])
def ai_compare():
    data = request.get_json()
    text1 = data.get('text1', '')
    text2 = data.get('text2', '')
    if not text1.strip() or not text2.strip():
        return jsonify({"error": "Both texts required"}), 400

    import difflib
    d = difflib.unified_diff(
        text1.splitlines(keepends=True),
        text2.splitlines(keepends=True),
        fromfile='Document 1',
        tofile='Document 2',
        n=2
    )
    diff_text = ''.join(list(d))

    # Use Mistral for summary of differences
    prompt = f"""Briefly summarize the key differences between these two documents in 3-4 bullet points:\n\nDocument 1 (first 500 chars):\n{text1[:500]}\n\nDocument 2 (first 500 chars):\n{text2[:500]}\n\nDifferences:"""
    summary = call_ollama(prompt, max_tokens=200)

    return jsonify({"success": True, "diff": diff_text, "summary": summary})


# ─── /api/export/excel ────────────────────────────────────────────────────────
@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    data = request.get_json()
    text = data.get('text', '')
    filename = data.get('filename', 'export')
    if not text.strip():
        return jsonify({"error": "No text provided"}), 400

    try:
        import openpyxl
        from flask import send_file
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Text"

        # Parse markdown tables
        lines = text.split('\n')
        row_idx = 1
        in_table = False
        for line in lines:
            if '|' in line and line.strip().startswith('|'):
                if '---' in line:
                    continue  # Skip separator rows
                cells = [c.strip() for c in line.strip().strip('|').split('|')]
                ws.append(cells)
                in_table = True
                row_idx += 1
            else:
                if in_table:
                    row_idx += 1  # blank row between tables
                    in_table = False
                if line.strip():
                    ws.cell(row=row_idx, column=1, value=line.strip())
                    row_idx += 1

        # Style header rows
        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="4472C4")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{filename}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── /api/ai/chat ─────────────────────────────────────────────────────────────
@app.route('/api/ai/chat', methods=['POST'])
def ai_chat():
    data = request.get_json()
    full_context = data.get('context', '')
    messages = data.get('messages', [])
    question = data.get('question', '')

    if not question.strip():
        return jsonify({"error": "Question required"}), 400

    context = full_context[:2000] # Default for small documents

    # --- RAG Logic for Large Documents ---
    rag_used = False
    if RAG_ENABLED and len(full_context) > 2000:
        try:
            # Simple chunking (1000 chars per chunk, 200 char overlap)
            chunk_size = 1000
            overlap = 200
            chunks = []
            for i in range(0, len(full_context), chunk_size - overlap):
                chunks.append(full_context[i:i+chunk_size])
            
            # Temporary collection for this query
            collection_name = f"doc_{hash(full_context[:100])}_{int(time.time())}"
            # Ensure safe collection name (alphanumeric, no hyphens, no negative signs)
            collection_name = collection_name.replace('-', '0')
            collection = chroma_client.create_collection(name=collection_name)
            
            # Create embeddings and add to collection
            embeddings = embedder.encode(chunks).tolist()
            collection.add(
                embeddings=embeddings,
                documents=chunks,
                ids=[str(i) for i in range(len(chunks))]
            )
            
            # Query the collection
            query_embedding = embedder.encode([question]).tolist()
            results = collection.query(
                query_embeddings=query_embedding,
                n_results=3
            )
            
            if results['documents'] and results['documents'][0]:
                retrieved_chunks = results['documents'][0]
                context = "\n...[RAG RETRIEVED CONTEXT]...\n".join(retrieved_chunks)
                rag_used = True
                
            # Cleanup
            chroma_client.delete_collection(name=collection_name)
        except Exception as e:
            print(f"RAG Error: {e}")
            # Fallback to default context truncation
            context = full_context[:3000]

    # Build conversation history
    history = ''
    for msg in messages[-4:]:  # Last 4 messages for context
        role = msg.get('role', 'user')
        content = msg.get('content', '')
        history += f"{role.upper()}: {content}\n"

    prompt = f"""You are a helpful document assistant. Answer questions based on the document below.
If the answer is not in the document, say so. Be concise.
If document is in Hindi, reply in Hindi.

Document Context:
{context}

{history}USER: {question}
ASSISTANT:"""

    answer = call_ollama(prompt, max_tokens=400)
    return jsonify({"success": True, "answer": answer, "rag_used": rag_used})


# ─── /api/health ──────────────────────────────────────────────────────────────
@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({
        "status": "healthy",
        "paddleocr_en": ocr_en is not None,
        "paddleocr_hi": ocr_hi is not None,
        "groq_model": GROQ_MODEL,
        "groq_ready": bool(GROQ_API_KEY)
    })

# ─── /api/magic-replace ────────────────────────────────────────────────────────
@app.route('/api/magic-replace', methods=['POST'])
def magic_replace():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    file = request.files['file']
    find_text = request.form.get('find_text', '')
    replace_text = request.form.get('replace_text', '')
    
    if file.filename == '' or not find_text:
        return jsonify({"error": "Missing file or find_text"}), 400

    try:
        from PIL import Image, ImageDraw, ImageFont
        temp_dir = tempfile.mkdtemp()
        filepath = os.path.join(temp_dir, file.filename)
        file.save(filepath)

        ocr_engine = ocr_en
        if ocr_engine is None:
            return jsonify({"error": "PaddleOCR is not initialized"}), 500

        result = ocr_engine.ocr(filepath)
        
        img = Image.open(filepath).convert("RGB")
        draw = ImageDraw.Draw(img)
        
        try:
            font = ImageFont.truetype("arial.ttf", 20)
        except:
            font = ImageFont.load_default()

        replaced_count = 0
        if result and result[0]:
            for line in result[0]:
                box = line[0]
                text = line[1][0]
                
                if find_text.lower() in text.lower():
                    xs = [p[0] for p in box]
                    ys = [p[1] for p in box]
                    x_min, x_max = min(xs), max(xs)
                    y_min, y_max = min(ys), max(ys)
                    
                    draw.rectangle([x_min-2, y_min-2, x_max+2, y_max+2], fill=(255, 255, 255))
                    
                    try:
                        box_height = int(y_max - y_min)
                        dyn_font = ImageFont.truetype("arial.ttf", max(10, box_height - 4))
                    except:
                        dyn_font = font
                        
                    draw.text((x_min, y_min), replace_text, fill=(0, 0, 0), font=dyn_font)
                    replaced_count += 1

        if replaced_count == 0:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return jsonify({"success": False, "message": "Text not found in image."}), 404

        output_path = os.path.join(temp_dir, f"replaced_{file.filename}")
        img.save(output_path)
        
        from flask import send_file
        return send_file(
            output_path,
            as_attachment=True,
            download_name=f"Replaced_{file.filename}",
            mimetype='image/png' if file.filename.lower().endswith('.png') else 'image/jpeg'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f"Starting TextMitra OCR Server on port {port}...")
    print(f"Groq AI : {GROQ_MODEL} ({'✓ Ready' if GROQ_API_KEY else '✗ No API key'})")
    app.run(port=port, host='0.0.0.0', debug=False)
